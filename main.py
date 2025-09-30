import streamlit as st
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
import time
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
import re

# Page configuration
st.set_page_config(
    page_title="EMA20 Scanner",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Modern UI styling with blue theme
st.markdown("""
<style>
    /* Overall page styling */
    .main {
        background-color: #f8f9fa;
        padding: 2rem;
    }
    
    /* Headers */
    h1 {
        color: #0d4b9f;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        font-weight: 700;
        font-size: 2.5rem;
        margin-bottom: 1.5rem;
    }
    
    h2, h3 {
        color: #334155;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        font-weight: 600;
    }
    
    /* Containers */
    .stApp {
        max-width: 1200px;
        margin: 0 auto;
    }
    
    /* Sidebar */
    section[data-testid="stSidebar"] {
        background-color: #ffffff;
        border-right: 1px solid #e2e8f0;
        padding: 2rem 1rem;
    }
    
    section[data-testid="stSidebar"] > div {
        padding-top: 0;
    }
    
    section[data-testid="stSidebar"] h2 {
        margin-top: 0;
    }
    
    /* Buttons */
    .stButton > button {
        background-color: #1e40af;
        color: white;
        border: none;
        border-radius: 6px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        transition: all 0.2s ease;
        width: 100%;
    }
    
    .stButton > button:hover {
        background-color: #1e3a8a;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
    }
    
    /* DataFrames */
    .dataframe {
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif !important;
    }
    
    .dataframe th {
        background-color: #f1f5f9;
        color: #334155;
        font-weight: 600;
        border: none !important;
        text-align: left !important;
    }
    
    .dataframe td {
        border-bottom: 1px solid #e2e8f0 !important;
        border-left: none !important;
        border-right: none !important;
        text-align: left !important;
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0;
        border-bottom: 1px solid #e2e8f0;
    }
    
    .stTabs [data-baseweb="tab"] {
        padding: 1rem 1.5rem;
        border: none;
        border-bottom: 2px solid transparent;
        font-weight: 500;
        color: #64748b;
        background-color: transparent;
    }
    
    .stTabs [aria-selected="true"] {
        border-bottom: 2px solid #1e40af !important;
        color: #1e40af !important;
        background-color: transparent !important;
    }
    
    /* Radio buttons */
    div[role="radiogroup"] label {
        background-color: white;
        border: 1px solid #e2e8f0;
        border-radius: 6px;
        padding: 0.5rem 1rem;
        margin-right: 0.5rem;
        font-weight: 500;
        transition: all 0.2s ease;
    }
    
    div[role="radiogroup"] label:hover {
        border-color: #cbd5e1;
        background-color: #f8fafc;
    }
    
    div[role="radiogroup"] [data-baseweb="radio"] input:checked + div {
        border-color: #2e7d32;
        background-color: #e8f5e9;
    }
    
    /* Select boxes */
    div[data-baseweb="select"] > div {
        border-radius: 6px !important;
        border-color: #e2e8f0 !important;
        background-color: white;
    }
    
    div[data-baseweb="select"] > div:hover {
        border-color: #cbd5e1 !important;
    }
    
    /* Info boxes */
    .stAlert {
        border-radius: 6px;
    }
    
    /* Fix for dark mode */
    @media (prefers-color-scheme: dark) {
        .stApp, body, [data-testid="stAppViewContainer"] {
            background-color: #0e1117;
        }
        
        h1, h2, h3, p, span, div {
            color: #f8f9fa;
        }
        
        .stTabs [data-baseweb="tab"] {
            color: #f8f9fa;
        }
        
        section[data-testid="stSidebar"] {
            background-color: #262730;
            border-right: 1px solid #4b5563;
        }
        
        .dataframe th {
            background-color: #1e293b;
            color: #f8f9fa;
        }
        
        .dataframe td {
            border-bottom: 1px solid #4b5563 !important;
            color: #f8f9fa;
        }
    }
</style>
""", unsafe_allow_html=True)

# Define stock markets data
us_indices = {
    'S&P 500': '^GSPC',
    'Dow Jones': '^DJI',
    'NASDAQ': '^IXIC'
}

india_indices = {
    'NIFTY 50': '^NSEI',
    'SENSEX': '^BSESN',
    'NIFTY BANK': '^NSEBANK'
}

# Function to sanitize symbols
def sanitize_symbol(symbol):
    """Sanitize stock symbols to prevent injection attacks"""
    if not isinstance(symbol, str):
        return ""
    
    # Allow only alphanumeric characters, dots, hyphens, and ^ for indices
    sanitized = re.sub(r'[^A-Za-z0-9.\-^]', '', str(symbol).strip())
    
    # Limit length to prevent abuse
    sanitized = sanitized[:20]
    
    return sanitized

# Function to sanitize company names
def sanitize_name(name):
    """Sanitize company names to prevent injection attacks"""
    if not isinstance(name, str):
        return ""
    
    # Allow alphanumeric, spaces, common punctuation
    sanitized = re.sub(r'[^A-Za-z0-9\s\.,&\-\(\)]', '', str(name).strip())
    
    # Limit length to prevent abuse
    sanitized = sanitized[:200]
    
    return sanitized

# Function to load stock lists
@st.cache_data(ttl=86400)
def load_stock_lists():
    # Load US Stocks from Excel
    try:
        us_stocks = pd.read_excel('data/us_stocks.xlsx')
        if not all(col in us_stocks.columns for col in ['Symbol', 'Company Name']):
            # Try alternative column names
            column_mapping = {}
            for col in us_stocks.columns:
                if col.lower() in ['symbol', 'ticker', 'stock']:
                    column_mapping[col] = 'Symbol'
                elif col.lower() in ['name', 'company', 'company name', 'stock name']:
                    column_mapping[col] = 'Company Name'
            
            if column_mapping:
                us_stocks = us_stocks.rename(columns=column_mapping)
    except Exception as e:
        st.warning(f"Failed to load US stocks Excel: {e}. Using default list.")
        us_stocks = pd.DataFrame({
            'Symbol': ['AAPL', 'MSFT', 'AMZN', 'GOOGL', 'META', 'TSLA', 'NVDA', 'JPM', 'V', 'WMT'],
            'Company Name': ['Apple', 'Microsoft', 'Amazon', 'Alphabet', 'Meta Platforms', 'Tesla', 'NVIDIA', 'JPMorgan Chase', 'Visa', 'Walmart']
        })
    
    # Load Indian Stocks from Excel
    try:
        india_stocks = pd.read_excel('data/india_stocks.xlsx')
        if not all(col in india_stocks.columns for col in ['Symbol', 'Company Name']):
            # Try alternative column names
            column_mapping = {}
            for col in india_stocks.columns:
                if col.lower() in ['symbol', 'ticker', 'stock']:
                    column_mapping[col] = 'Symbol'
                elif col.lower() in ['name', 'company', 'company name', 'stock name']:
                    column_mapping[col] = 'Company Name'
            
            if column_mapping:
                india_stocks = india_stocks.rename(columns=column_mapping)
        
        # Ensure Indian stock symbols have .NS suffix for API calls
        india_stocks['Symbol'] = india_stocks['Symbol'].apply(
            lambda x: sanitize_symbol(x) if str(x).endswith('.NS') else f"{sanitize_symbol(x)}.NS"
        )
    except Exception as e:
        st.warning(f"Failed to load India stocks Excel: {e}. Using default list.")
        india_stocks = pd.DataFrame({
            'Symbol': ['RELIANCE.NS', 'TCS.NS', 'HDFCBANK.NS', 'INFY.NS', 'ICICIBANK.NS', 
                     'HINDUNILVR.NS', 'ITC.NS', 'SBIN.NS', 'BAJFINANCE.NS', 'BHARTIARTL.NS'],
            'Company Name': ['Reliance Industries', 'Tata Consultancy Services', 'HDFC Bank', 'Infosys', 
                    'ICICI Bank', 'Hindustan Unilever', 'ITC', 'State Bank of India', 
                    'Bajaj Finance', 'Bharti Airtel']
        })
    
    # Sanitize all symbols and names
    us_stocks['Symbol'] = us_stocks['Symbol'].apply(sanitize_symbol)
    us_stocks['Company Name'] = us_stocks['Company Name'].apply(sanitize_name)
    india_stocks['Symbol'] = india_stocks['Symbol'].apply(sanitize_symbol)
    india_stocks['Company Name'] = india_stocks['Company Name'].apply(sanitize_name)
    
    # Remove empty entries
    us_stocks = us_stocks[(us_stocks['Symbol'].str.len() > 0) & (us_stocks['Company Name'].str.len() > 0)]
    india_stocks = india_stocks[(india_stocks['Symbol'].str.len() > 0) & (india_stocks['Company Name'].str.len() > 0)]
    
    return us_stocks, india_stocks

# Function to process uploaded stock list
def process_uploaded_stock_list(uploaded_file, market):
    try:
        # Read Excel file only
        if not uploaded_file.name.endswith('.xlsx'):
            st.error("Only Excel (.xlsx) files are supported")
            return None
            
        stocks_df = pd.read_excel(uploaded_file)
        
        # Standardize column names (case-insensitive)
        column_mapping = {}
        for col in stocks_df.columns:
            if col.lower() in ['symbol', 'ticker', 'stock']:
                column_mapping[col] = 'Symbol'
            elif col.lower() in ['name', 'company', 'company name', 'stock name']:
                column_mapping[col] = 'Company Name'
        
        # Rename columns if needed
        if column_mapping:
            stocks_df = stocks_df.rename(columns=column_mapping)
        
        # Check if we have the required columns
        if 'Symbol' not in stocks_df.columns:
            raise ValueError("File must contain a 'Symbol' column")
        
        # If no name column exists, create one with symbol values
        if 'Company Name' not in stocks_df.columns:
            stocks_df['Company Name'] = stocks_df['Symbol']
        
        # Sanitize all data
        stocks_df['Symbol'] = stocks_df['Symbol'].apply(sanitize_symbol)
        stocks_df['Company Name'] = stocks_df['Company Name'].apply(sanitize_name)
        
        # Remove empty entries
        stocks_df = stocks_df[(stocks_df['Symbol'].str.len() > 0) & (stocks_df['Company Name'].str.len() > 0)]
        
        # Ensure proper formatting for Indian stocks (but not for indices starting with ^)
        if market == "India":
            stocks_df['Symbol'] = stocks_df['Symbol'].apply(
                lambda x: x if str(x).startswith('^') or str(x).endswith('.NS') else f"{x}.NS"
            )
        
        # Limit to 9999 stocks
        if len(stocks_df) > 9999:
            stocks_df = stocks_df.iloc[:9999]
            st.warning(f"Stock list limited to 9999 stocks")
        
        return stocks_df
        
    except Exception as e:
        st.error(f"Error processing uploaded file: {e}")
        return None

# Function to get stock data and calculate EMA20
@st.cache_data(ttl=3600)
def get_stock_data(symbol, timeframe):
    try:
        # Sanitize symbol before API call
        symbol = sanitize_symbol(symbol)
        if not symbol:
            return None
            
        stock = yf.Ticker(symbol)
        
        # Set period and interval based on timeframe
        timeframe_config = {
            "1h": {"period": "5d", "interval": "1h"},      # Hourly -> 5 Days
            "1d": {"period": "40d", "interval": "1d"},     # Daily -> 40 days
            "1wk": {"period": "6mo", "interval": "1wk"},   # Weekly -> 6 Months
            "1mo": {"period": "2y", "interval": "1mo"}     # Monthly -> 2 Years
        }
        
        config = timeframe_config.get(timeframe, timeframe_config["1d"])
        
        df = stock.history(period=config["period"], interval=config["interval"])
        
        if df.empty or len(df) < 20:  # Ensure we have enough data for EMA20
            return None
        
        # Calculate EMA20 precisely using exponential weighting
        df['EMA20'] = df['Close'].ewm(span=20, adjust=False).mean()
        
        return df
    except Exception as e:
        return None

# Function to check EMA20 position vs current candle close
def check_ema20_position(df):
    if df is None or df.empty:
        return None, None
    
    # Get the latest closed candle values
    latest = df.iloc[-1]
    close_price = latest['Close']
    ema20 = latest['EMA20']
    
    # Check position relative to EMA20
    if close_price > ema20:
        return "Bullish", "🟢"
    elif close_price < ema20:
        return "Bearish", "🔴"
    else:
        return None, None

# Function to scan all stocks for EMA20 position
def scan_ema20_position(stock_list, timeframe, market):
    results = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_stocks = len(stock_list)
    processed_count = 0
    
    for i, (symbol, name) in enumerate(zip(stock_list['Symbol'], stock_list['Company Name'])):
        status_text.text(f"Scanning {market} stocks: {i+1}/{total_stocks} - {name} ({symbol})")
        progress_bar.progress((i + 1) / total_stocks)
        
        df = get_stock_data(symbol, timeframe)
        
        if df is None or df.empty:
            continue
            
        processed_count += 1
        
        trend, status_emoji = check_ema20_position(df)
        
        if trend:  # Add both bullish and bearish stocks
            # Get current price and EMA20 for additional info
            latest = df.iloc[-1]
            current_price = latest['Close']
            ema20_value = latest['EMA20']
            
            # Remove .NS suffix and ^ symbol for display
            display_symbol = symbol.replace('.NS', '') if symbol.endswith('.NS') else symbol
            display_symbol = display_symbol.replace('^', '') if display_symbol.startswith('^') else display_symbol
            
            results.append({
                'Symbol': display_symbol,
                'Company Name': name,
                'Trend': trend,
                'Current Price': f"{current_price:.2f}",
                'EMA20': f"{ema20_value:.2f}",
                'Original_Symbol': symbol  # Keep original for any further processing
            })
    
    progress_bar.empty()
    status_text.empty()
    
    # Show summary of scan results
    if processed_count < total_stocks:
        st.info(f"Note: Data for {total_stocks - processed_count} stocks could not be retrieved or processed.")
    
    return pd.DataFrame(results) if results else pd.DataFrame()

# Function to create formatted Excel file
def create_formatted_excel(df, filename):
    if df.empty:
        return None
    
    # Create a copy of dataframe for export (without Original_Symbol)
    export_df = df[['Symbol', 'Company Name', 'Trend', 'Current Price', 'EMA20']].copy()
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    try:
        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'EMA20 Scanner Results'
        
        # Write headers
        headers = ['Symbol', 'Company Name', 'Trend', 'Current Price', 'EMA20']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Define colors and fills
        green_font = Font(color="00008000", bold=True)  # Green
        red_font = Font(color="00FF0000", bold=True)    # Red
        green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Light green background
        red_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")    # Light red background
        
        # Write data rows
        for row_num, (_, row_data) in enumerate(export_df.iterrows(), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                
                # Apply formatting based on trend
                if row_data['Trend'] == 'Bullish':
                    cell.font = green_font
                    cell.fill = green_fill
                elif row_data['Trend'] == 'Bearish':
                    cell.font = red_font
                    cell.fill = red_fill
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        workbook.save(output)
        output.seek(0)
        
        return output
    
    except Exception as e:
        st.error(f"Error creating Excel file: {e}")
        return None

# Main application
def main():
    st.title("EMA20 Scanner")
    
    # Display current market status at the top
    st.subheader("Market Status")
    col1, col2, col3 = st.columns(3)
    
    # Initialize session state for managing stock lists
    if 'using_custom_list' not in st.session_state:
        st.session_state.using_custom_list = False
    
    # Sidebar
    st.sidebar.header("Scanner Settings")
    
    # Load default stock lists
    us_stocks, india_stocks = load_stock_lists()
    
    # Custom stock list upload
    st.sidebar.subheader("Stock List")
    uploaded_file = st.sidebar.file_uploader(
        "Upload Custom (Symbol, Company Name)",
        type=["xlsx"],
        help="Excel file with 'Symbol' and 'Company Name' columns (Max 50MB, 9999 stocks)"
    )
    
    # Process uploaded file if available
    custom_stocks = None
    if uploaded_file is not None:
        if uploaded_file.size > 50 * 1024 * 1024:  # 50MB limit
            st.sidebar.error("File size exceeds 50MB limit")
            st.session_state.using_custom_list = False
        else:
            market_for_processing = st.session_state.get('market', "India")
            custom_stocks = process_uploaded_stock_list(uploaded_file, market_for_processing)
            
            if custom_stocks is not None:
                st.session_state.using_custom_list = True
                st.session_state.custom_stocks = custom_stocks
                st.sidebar.success(f"Loaded {len(custom_stocks)} stocks from your file")
            else:
                st.session_state.using_custom_list = False
    else:
        st.session_state.using_custom_list = False
    
    # Market selection - Default to India
    if st.session_state.using_custom_list:
        market = st.sidebar.selectbox(
            "Select Market (Disabled - Using Custom List)",
            ["India", "US"],
            disabled=True,
            index=0 if st.session_state.get('market', "India") == "India" else 1
        )
        market = st.session_state.get('market', "India")
    else:
        market = st.sidebar.selectbox("Select Market", ["India", "US"])
        st.session_state.market = market
    
    # Timeframe selection with new options
    timeframe_options = {
        "Hourly (5 Days)": "1h",
        "Daily (40 Days)": "1d", 
        "Weekly (6 Months)": "1wk",
        "Monthly (2 Years)": "1mo"
    }
    timeframe_display = st.sidebar.selectbox("Select Timeframe", list(timeframe_options.keys()), index=1)  # Default to Daily
    timeframe = timeframe_options[timeframe_display]
    
    # Scan button
    scan_button = st.sidebar.button("Start EMA20 Scanner", use_container_width=True)
    
    # Display current market status data
    indices = india_indices if market == "India" else us_indices
    
    index_cols = [col1, col2, col3]
    for i, (index_name, index_symbol) in enumerate(indices.items()):
        try:
            # Sanitize index symbol
            sanitized_index_symbol = sanitize_symbol(index_symbol)
            index_data = yf.Ticker(sanitized_index_symbol).history(period="1d")
            if not index_data.empty:
                current = index_data['Close'].iloc[-1]
                previous = index_data['Open'].iloc[-1]
                change = current - previous
                change_percent = (change / previous) * 100
                
                color = "green" if change >= 0 else "red"
                change_icon = "▲" if change >= 0 else "▼"
                
                index_cols[i].markdown(
                    f"**{index_name}**: {current:.2f} "
                    f"<span style='color:{color}'>{change_icon} {abs(change):.2f} ({abs(change_percent):.2f}%)</span>", 
                    unsafe_allow_html=True
                )
        except:
            index_cols[i].text(f"{index_name}: Data unavailable")
    
    if scan_button:
        # Use custom stock list if uploaded, otherwise use default
        if st.session_state.using_custom_list:
            stocks_to_scan = st.session_state.custom_stocks
        else:
            stocks_to_scan = india_stocks if market == "India" else us_stocks
        
        with st.spinner(f"Scanning {market} stocks for EMA20 position on {timeframe_display}..."):
            results_df = scan_ema20_position(stocks_to_scan, timeframe, market)
        
        # Store results in session state
        st.session_state.results_df = results_df
        st.session_state.last_scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        st.session_state.market = market
        st.session_state.timeframe = timeframe_display
    
    # Display explanation
    with st.expander("How This EMA20 Scanner Works"):
        st.markdown(f"""
        **Disclaimer**: This project is intended for educational and informational purposes only. You are solely responsible for any profits or losses you may incur.
        
        ## EMA20 Scanner Logic
        
        This scanner identifies stocks based on their position relative to the 20-period EMA (Exponential Moving Average):
        
        ### 🟢 Bullish Stocks
        - Current candle close price is **ABOVE** the 20 EMA
        - Indicates potential upward momentum
        
        ### 🔴 Bearish Stocks  
        - Current candle close price is **BELOW** the 20 EMA
        - Indicates potential downward momentum
        
        ### Timeframes and Data Periods
        - **Hourly (5 Days)**: Uses 5 days of hourly data for short-term analysis
        - **Daily (40 Days)**: Uses 40 days of daily data for medium-term analysis  
        - **Weekly (6 Months)**: Uses 6 months of weekly data for swing analysis
        - **Monthly (2 Years)**: Uses 2 years of monthly data for long-term analysis
        
        ### Important Notes
        - **Closed Candle Basis**: All calculations are performed only on completed/closed candles
        - EMA20 is calculated using exponential weighting (span=20)
        - Indian stock symbols display without .NS suffix in results
        - Export files are formatted with color coding (Green for Bullish, Red for Bearish)
        - All data is sanitized for security
        
        ### Using Custom Stock Lists
        - Upload Excel files with 'Symbol' and 'Company Name' columns
        - Maximum 9999 stocks per list and 50MB file size
        - For Indian stocks, .NS suffix is automatically handled
        - For indices starting with ^, .NS suffix is not added
        """)
    
    # Display results
    if 'results_df' in st.session_state and not st.session_state.results_df.empty:
        st.subheader("EMA20 Scanner Results")
        
        # Show last scan info
        scan_info = f"Last scan: {st.session_state.last_scan_time} | Market: {st.session_state.market} | Timeframe: {st.session_state.timeframe}"
        st.info(scan_info)
        
        # Separate bullish and bearish stocks
        bullish_stocks = st.session_state.results_df[st.session_state.results_df['Trend'] == 'Bullish']
        bearish_stocks = st.session_state.results_df[st.session_state.results_df['Trend'] == 'Bearish']
        
        # Create tabs for bullish and bearish
        tab1, tab2 = st.tabs([f"Bullish Stocks 🟢 ({len(bullish_stocks)})", f"Bearish Stocks 🔴 ({len(bearish_stocks)})"])
        
        with tab1:
            if not bullish_stocks.empty:
                st.subheader("Stocks Trading Above EMA20")
                display_df = bullish_stocks[['Symbol', 'Company Name', 'Current Price', 'EMA20']].copy()
                st.dataframe(display_df, use_container_width=True)
                
                # Download button for bullish stocks - Excel format
                excel_file = create_formatted_excel(bullish_stocks, f"bullish_stocks_{st.session_state.market}_{st.session_state.timeframe}")
                if excel_file:
                    st.download_button(
                        label="📥 Download Bullish Stocks (Excel)",
                        data=excel_file.getvalue(),
                        file_name=f"bullish_ema20_{st.session_state.market}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_bullish"
                    )
            else:
                st.info("No stocks found trading above EMA20.")
        
        with tab2:
            if not bearish_stocks.empty:
                st.subheader("Stocks Trading Below EMA20")
                display_df = bearish_stocks[['Symbol', 'Company Name', 'Current Price', 'EMA20']].copy()
                st.dataframe(display_df, use_container_width=True)
                
                # Download button for bearish stocks - Excel format
                excel_file = create_formatted_excel(bearish_stocks, f"bearish_stocks_{st.session_state.market}_{st.session_state.timeframe}")
                if excel_file:
                    st.download_button(
                        label="📥 Download Bearish Stocks (Excel)",
                        data=excel_file.getvalue(),
                        file_name=f"bearish_ema20_{st.session_state.market}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_bearish"
                    )
            else:
                st.info("No stocks found trading below EMA20.")
        
        # Download all results button - Excel format
        if not st.session_state.results_df.empty:
            st.subheader("Download All Results")
            excel_file = create_formatted_excel(st.session_state.results_df, f"ema20_results_{st.session_state.market}_{st.session_state.timeframe}")
            if excel_file:
                st.download_button(
                    label="📥 Download All Results (Excel)",
                    data=excel_file.getvalue(),
                    file_name=f"ema20_all_results_{st.session_state.market}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_all"
                )
    
    elif 'last_scan_time' in st.session_state:
        st.info("No stocks found with clear EMA20 positioning. Try scanning with different parameters.")
    else:
        st.info("Click 'Start EMA20 Scanner' to begin scanning for stocks above or below EMA20.")

# Run the application
if __name__ == "__main__":
    main()